from django.shortcuts import render
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.views import View
from django.views.generic import TemplateView
from .models import Robot
from .serializers import RobotSerializer
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime


# Create your views here.
class RobotCreateView(View):
    def post(self, request):
        robot = RobotSerializer()
        robot.parse_json(request)
        if robot.is_valid():
            robot.save()
            response_data = {
                'message': 'Robot created successfully!',
                'robot_data': {
                    'serial': robot.serial,
                    'model': robot.model,
                    'version': robot.version,
                    'created': robot.created.strftime("%Y-%m-%d %H:%M:%S")
                }
            }
            return JsonResponse(response_data, status=201)
        else:
            return JsonResponse({
                'error': 'Invalid JSON format or data. Valid format: {"model":"model_name","version":"version_name","created":"date"}, format date:"%Y-%m-%d %H:%M:%S"'},
                status=400)


class RobotReportView(TemplateView):
    template_name = 'robots_week_report.html'
    TIMEDELTA = 7

    @staticmethod
    def apply_style_wb(ws, count_col, header_font, header_fill):
        """Применяет стили к заголовкам"""
        for col_num in range(1, count_col + 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}1"]
            cell.font = header_font
            cell.fill = header_fill

    def download_report(request):

        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")

        wb = openpyxl.Workbook()
        ws = wb.active
        end_date = datetime.datetime.now()
        start_date = end_date - datetime.timedelta(days=RobotReportView.TIMEDELTA)

        robots = Robot.objects.filter(created__gte=start_date, created__lte=end_date)

        # Создаем страницу "Сводка"
        ws.title = "Сводка"
        ws['A1'] = "Модель"
        ws['B1'] = "Версия"
        ws['C1'] = "Количество"

        RobotReportView.apply_style_wb(ws, 3, header_font, header_fill)

        model_data = {}
        for robot in robots:
            model = robot.model
            version = robot.version
            if model not in model_data:
                model_data[model] = {}
            if version not in model_data[model]:
                model_data[model][version] = 0
            model_data[model][version] += 1

        # Заполняем страницу "Сводка"
        row = 2
        for model, version_data in model_data.items():
            for version, count in version_data.items():
                ws.cell(row=row, column=1, value=model)
                ws.cell(row=row, column=2, value=version)
                ws.cell(row=row, column=3, value=count)
                row += 1

        # Создаем страницы для каждой модели
        models = Robot.objects.values_list('model', flat=True).distinct()

        for model in models:
            model_ws = wb.create_sheet(title=model)
            model_ws['A1'] = "Модель"
            model_ws['B1'] = "Версия"
            model_ws['C1'] = "Дата создания"

            RobotReportView.apply_style_wb(model_ws, 3, header_font, header_fill)

            model_robots = Robot.objects.filter(model=model, created__gte=start_date, created__lte=end_date)

            row = 2
            for robot in model_robots:
                model_ws.cell(row=row, column=1, value=robot.model)
                model_ws.cell(row=row, column=2, value=robot.version)
                model_ws.cell(row=row, column=3, value=str(robot.created))
                row += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=robot_production_summary.xlsx'
        wb.save(response)

        return response
